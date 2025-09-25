from get_information import get_summarized_area, input_workarea, get_actual_bote
from get_information import text_day, date_text, cantidad_ex, num_partida, equip
from get_information import total_haz_waste_quantity, total_safe_waste_quantity
from get_information import safe_waste_data, dangerous_waste_data
from openpyxl.drawing.image import Image
import os

def diary_format(sheet_diary):
    if num_partida and cantidad_ex:

        print(f'Working in the item N# {num_partida}')

        weekdays = {'LUNES': 0, 'MARTES': 1, 'MIERCOLES': 2, 'JUEVES': 3, 'VIERNES': 4, 'SABADO': 5}
        columns = {0: 9, 1: 10, 2: 11, 3: 12, 4: 13, 5: 7}

        for row in sheet_diary.iter_rows(min_row=11, max_row=26, min_col=1, max_col=1):
            for cell in row:
                if str(cell.value) == num_partida:
                    sheet_diary.cell(row=cell.row, column=columns[weekdays[text_day]], value=cantidad_ex)
                    break

        return f"Diary format for {text_day} done!"
    raise ValueError("There is no Valid item (partida) Number or quantity in this inform")

def put_images(sheet):
    images_added = False
    
    for filename in os.listdir('Inform_images'):
        if filename.endswith(('.png', '.jpg', '.jpeg')):
            img_path = os.path.join('Inform_images', filename)
            img = Image(img_path)
            img.height = 193.914
            img.width = 119.826
            sheet.add_image(img, 'B34')
            images_added = True
    
    if not images_added:
        raise ValueError('There are no valid images in the Images folder!')
    return f"Images putted in {text_day}"

def put_date(sheet):
    if date_text and sheet:
        sheet.cell(row=6, column=9, value=date_text)
        return f"Date put in {text_day}"
    raise ValueError('There is no valid Date in this inform')

def put_area(sheet):
    add_text, sum_area = input_workarea()
    sheet.cell(row=6, column=1, value= get_summarized_area(sum_area))

    for i, work_area in enumerate(add_text):
        sheet.cell(row=26 + i, column=2, value=work_area)
        sheet.cell(row=26 + i, column=1, value=1 + i)
    
    return f"Área put in {text_day}"

def clean_area(sheet):
    for i in range(7):
        sheet.cell(row=26 + i, column= 1, value='')
        sheet.cell(row=26 + i, column= 2, value='')
    return f"{text_day} Área cleaned!"

def put_equipment(sheet):
    sheet.cell(row=26, column=7, value=equip)
    sheet.cell(row=26, column=9, value=8)
    return f"Equipment put in {text_day}"

"""
def put_personal(sheet):
    for i, person in enumerate(indirect_personal):
        sheet.cell(row=11 + i, column=1, value=person)
        sheet.cell(row=11 + i, column=4, value=1)
        
    # Adding the Direct personal
    sheet.cell(row=11, column= 5, value= list(direct_workers.keys())[0])
    sheet.cell(row=11, column = 6, value= list(direct_workers.values())[0])

    for j, (key, value) in enumerate(indirect_workers.items()):
        sheet.cell(row=11 + j, column = 7, value= key)
        sheet.cell(row=11 + j, column = 9, value = value)

    return f"Personal put in {text_day}!"
"""


def put_bote(sheet):
    total_last_safe_quantity, total_last_dangerous_quantity = get_actual_bote(sheet)
    
    if total_haz_waste_quantity == total_last_dangerous_quantity and total_safe_waste_quantity == total_last_safe_quantity:
        return f"Nothing has changed in the Bote for {text_day}!"
    
    if total_safe_waste_quantity != total_last_safe_quantity:
        print(put_safe_bote(sheet, safe_waste_data, total_last_safe_quantity))
    
    if total_haz_waste_quantity != total_last_dangerous_quantity:
       print(put_dangerous_bote(sheet, dangerous_waste_data, total_last_dangerous_quantity))
    
    put_equipment(sheet)
    put_bote_images(sheet)
        
    return f"The BOTE dates and quantities has been updated!"

def put_bote_images(sheet):
    images_added = False
    
    for file in os.listdir('Bote_images'):
        if file.endswith(('.png', '.jpg', '.jpeg')):
            img_path = os.path.join('Bote_images', file)
            img = Image(img_path)
            img.height = 208.287
            img.width = 287.28
            sheet.add_image(img, 'B35')
            images_added = True
    
    if not images_added:
        print("There are no images in the Bote folder!")

    return f"Images added on {text_day}"

def put_dangerous_bote(sheet, dangerous_waste_data, total_last_dangerous_quantity):
    act_dangerous_date, act_dangerous_quantity = dangerous_waste_data[-1]
    act_dangerous_date = act_dangerous_date.replace("-","/")
    
    sheet.cell(row= 32, column= 1, value= act_dangerous_date)
    sheet.cell(row= 32, column= 4, value= act_dangerous_date)
    sheet.cell(row= 6, column= 6, value= act_dangerous_date)

    sheet.cell(row= 32, column= 3, value= float(act_dangerous_quantity))
    sheet.cell(row= 29, column= 1, value= total_last_dangerous_quantity)
    return "Dangerous data in the BOTE has been updated!"

def put_safe_bote(sheet, safe_waste_data, total_last_safe_quantity):
    act_safe_date, act_safe_quantity = safe_waste_data[-1]
    act_safe_date = act_safe_date.replace("-","/")
    
    sheet.cell(row= 32, column= 4, value= act_safe_date)
    sheet.cell(row= 32, column= 1, value= act_safe_date)
    sheet.cell(row= 6, column= 6, value= act_safe_date)

    sheet.cell(row= 32, column= 5, value= float(act_safe_quantity))
    sheet.cell(row= 29, column= 4, value= total_last_safe_quantity)
    return "Safe data in the BOTE has been updated!"